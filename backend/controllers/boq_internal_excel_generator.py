"""
Internal BOQ Excel Generator
Generates detailed internal BOQ with materials, labour, and cost breakdowns
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import date


def generate_internal_excel(project, items, total_material_cost, total_labour_cost, grand_total, boq_json=None):
    """
    Generate Internal Excel file with DETAILED cost breakdown
    Shows materials, labour, internal costs, and profit analysis
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Internal BOQ"

    if boq_json is None:
        boq_json = {}

    # Define styles
    header_font = Font(bold=True, size=16, color="1F4788")
    sub_header_font = Font(bold=True, size=12, color="1F4788")
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=9)
    small_font = Font(size=8, color="666666")

    # Header fills
    blue_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    grey_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "INTERNAL BOQ - Bill of Quantities"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "Confidential - Internal Use Only"
    ws[f'A{row}'].font = Font(size=10, italic=True, color="EF4444")
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Project Information
    ws[f'A{row}'] = "Project Information"
    ws[f'A{row}'].font = sub_header_font
    row += 1

    info_data = [
        ["Project Name:", project.project_name],
        ["Client:", project.client],
        ["Location:", project.location],
        ["Date:", date.today().strftime('%d %B %Y')]
    ]

    for info_row in info_data:
        ws[f'A{row}'] = info_row[0]
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = grey_fill
        ws[f'B{row}'] = info_row[1]
        ws[f'B{row}'].font = normal_font
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    row += 2

    # Preliminaries
    preliminaries = boq_json.get('preliminaries', {})
    prelim_items = preliminaries.get('items', [])
    prelim_notes = preliminaries.get('notes', '')

    if prelim_items or prelim_notes:
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "📋 PRELIMINARIES & APPROVAL WORKS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="92400E")
        ws[f'A{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Show all preliminary items (already filtered to selected only)
        for idx, prelim_item in enumerate(prelim_items, 1):
            if isinstance(prelim_item, dict):
                desc = prelim_item.get('description', prelim_item.get('name', prelim_item.get('text', '')))
            else:
                desc = str(prelim_item)

            if desc:
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = f"{idx}. {desc}"
                ws[f'A{row}'].font = normal_font
                ws[f'A{row}'].fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
                ws[f'A{row}'].border = thin_border

                # Add custom badge if applicable
                if prelim_item.get('isCustom'):
                    ws[f'G{row}'] = "Custom"
                    ws[f'G{row}'].font = Font(size=8, color="D97706", italic=True)
                    ws[f'G{row}'].border = thin_border
                row += 1

        # Cost Summary
        cost_details = preliminaries.get('cost_details', {})
        if cost_details and cost_details.get('amount'):
            row += 1
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "Cost Summary"
            ws[f'A{row}'].font = Font(bold=True, size=10, color="92400E")
            ws[f'A{row}'].fill = yellow_fill
            row += 1

            cost_info = [
                ["Quantity:", cost_details.get('quantity', 1)],
                ["Unit:", cost_details.get('unit', 'Nos')],
                ["Rate:", f"{cost_details.get('rate', 0):,.2f} AED"],
                ["Total Amount:", f"{cost_details.get('amount', 0):,.2f} AED"]
            ]

            for cost_row in cost_info:
                ws[f'A{row}'] = cost_row[0]
                ws[f'A{row}'].font = bold_font
                ws[f'A{row}'].fill = yellow_fill
                ws[f'A{row}'].border = thin_border
                ws[f'B{row}'] = cost_row[1]
                ws[f'B{row}'].font = normal_font if cost_row[0] != "Total Amount:" else bold_font
                ws[f'B{row}'].border = thin_border
                row += 1

        if prelim_notes:
            row += 1
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = f"📝 Note: {prelim_notes}"
            ws[f'A{row}'].font = Font(italic=True, size=9, color="78350F")
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws[f'A{row}'].fill = yellow_fill
            ws[f'A{row}'].border = thin_border
            row += 1

        row += 2

    # Items breakdown
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "DETAILED COST BREAKDOWN"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Process each item
    for idx, item in enumerate(items, 1):
        # Item Header
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"{idx}. {item.get('item_name', 'N/A')}"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4788")
        ws[f'A{row}'].fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        row += 1

        if item.get('description'):
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = item['description']
            ws[f'A{row}'].font = Font(italic=True, size=9, color="6B7280")
            row += 1

        row += 1

        # Check if item has sub-items
        has_sub_items = item.get('has_sub_items', False)
        sub_items = item.get('sub_items', [])

        if has_sub_items and sub_items:
            for sub_idx, sub_item in enumerate(sub_items, 1):
                # Sub-item header with client amount
                qty = sub_item.get('quantity', 0)
                rate = sub_item.get('rate', 0)
                client_amount = qty * rate

                sub_item_header_row = row
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = f"{idx}.{sub_idx} {sub_item.get('sub_item_name', 'N/A')}"
                ws[f'A{row}'].font = Font(bold=True, size=10)
                ws[f'F{row}'] = "Client Amount:"
                ws[f'F{row}'].font = bold_font
                ws[f'F{row}'].alignment = Alignment(horizontal='right')
                ws[f'G{row}'] = round(client_amount, 2)
                ws[f'G{row}'].font = Font(bold=True, color="16A34A")
                ws[f'G{row}'].alignment = Alignment(horizontal='right')
                ws[f'G{row}'].number_format = '#,##0.00'
                row += 1

                # Sub-item details (scope, size, location, brand, qty, rate)
                details = []
                if sub_item.get('scope'):
                    details.append(f"Scope: {sub_item['scope']}")
                if sub_item.get('size'):
                    details.append(f"Size: {sub_item['size']}")
                if sub_item.get('location'):
                    details.append(f"Location: {sub_item['location']}")
                if sub_item.get('brand'):
                    details.append(f"Brand: {sub_item['brand']}")

                if details:
                    ws.merge_cells(f'A{row}:G{row}')
                    ws[f'A{row}'] = " | ".join(details)
                    ws[f'A{row}'].font = small_font
                    row += 1

                ws.merge_cells(f'A{row}:G{row}')
                ws[f'A{row}'] = f"Qty: {qty} {sub_item.get('unit', 'nos')} @ AED{rate:.2f}/{sub_item.get('unit', 'nos')}"
                ws[f'A{row}'].font = small_font
                row += 1

                row += 1

                # Materials section
                materials = sub_item.get('materials', [])
                if materials:
                    ws.merge_cells(f'A{row}:G{row}')
                    ws[f'A{row}'] = "+ RAW MATERIALS"
                    ws[f'A{row}'].font = Font(bold=True, size=9, color="FFFFFF")
                    ws[f'A{row}'].fill = light_red_fill
                    row += 1

                    # Materials table header
                    mat_headers = ['Material Name', 'Qty', 'Unit', 'Unit Price', 'Total']
                    for col_idx, header in enumerate(mat_headers, start=2):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = header
                        cell.font = Font(bold=True, size=9)
                        cell.fill = grey_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border
                    row += 1

                    materials_cost = 0
                    for mat in materials:
                        mat_total = mat.get('total_price', 0)
                        materials_cost += mat_total

                        ws.cell(row=row, column=2).value = mat.get('material_name', 'N/A')
                        ws.cell(row=row, column=3).value = round(mat.get('quantity', 0), 2)
                        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=4).value = mat.get('unit', 'nos')
                        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=5).value = round(mat.get('unit_price', 0), 2)
                        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                        ws.cell(row=row, column=5).number_format = '#,##0.00'
                        ws.cell(row=row, column=6).value = round(mat_total, 2)
                        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                        ws.cell(row=row, column=6).number_format = '#,##0.00'

                        for col in range(2, 7):
                            ws.cell(row=row, column=col).border = thin_border
                            ws.cell(row=row, column=col).font = normal_font

                        row += 1

                    # Total Materials
                    ws.cell(row=row, column=5).value = "Total Materials:"
                    ws.cell(row=row, column=5).font = bold_font
                    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                    ws.cell(row=row, column=6).value = round(materials_cost, 2)
                    ws.cell(row=row, column=6).font = bold_font
                    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                    row += 2

                # Labour section
                labour = sub_item.get('labour', [])
                if labour:
                    ws.merge_cells(f'A{row}:G{row}')
                    ws[f'A{row}'] = "+ LABOUR"
                    ws[f'A{row}'].font = Font(bold=True, size=9, color="FFFFFF")
                    ws[f'A{row}'].fill = light_red_fill
                    row += 1

                    # Labour table header
                    lab_headers = ['Labour Role', 'Hours', 'Unit', 'Rate/Hour', 'Total']
                    for col_idx, header in enumerate(lab_headers, start=2):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = header
                        cell.font = Font(bold=True, size=9)
                        cell.fill = grey_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border
                    row += 1

                    labour_cost = 0
                    for lab in labour:
                        lab_total = lab.get('total_cost', 0)
                        labour_cost += lab_total

                        ws.cell(row=row, column=2).value = f"{lab.get('labour_role', 'N/A')} (Labour)"
                        ws.cell(row=row, column=3).value = round(lab.get('hours', 0), 2)
                        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=4).value = "Hrs"
                        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=5).value = round(lab.get('rate_per_hour', 0), 2)
                        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                        ws.cell(row=row, column=5).number_format = '#,##0.00'
                        ws.cell(row=row, column=6).value = round(lab_total, 2)
                        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                        ws.cell(row=row, column=6).number_format = '#,##0.00'

                        for col in range(2, 7):
                            ws.cell(row=row, column=col).border = thin_border
                            ws.cell(row=row, column=col).font = normal_font

                        row += 1

                    # Total Labour
                    ws.cell(row=row, column=5).value = "Total Labour:"
                    ws.cell(row=row, column=5).font = bold_font
                    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                    ws.cell(row=row, column=6).value = round(labour_cost, 2)
                    ws.cell(row=row, column=6).font = bold_font
                    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                    row += 2

                # Calculations - Get percentages (same as PDF)
                materials_cost_calc = sum([m.get('total_price', 0) for m in sub_item.get('materials', [])])
                labour_cost_calc = sum([l.get('total_cost', 0) for l in sub_item.get('labour', [])])
                base_cost = materials_cost_calc + labour_cost_calc

                # Get percentages from SUB-ITEM first, fallback to PARENT ITEM
                misc_pct = sub_item.get('misc_percentage', item.get('miscellaneous_percentage', item.get('overhead_percentage', 10)))
                overhead_pct = sub_item.get('overhead_profit_percentage', item.get('overhead_profit_percentage', item.get('profit_margin_percentage', 25)))
                transport_pct = sub_item.get('transport_percentage', item.get('transport_percentage', 5))

                # Calculate based on CLIENT AMOUNT (as per BOQDetailsModal)
                misc_amt = client_amount * (misc_pct / 100)
                overhead_amt = client_amount * (overhead_pct / 100)
                transport_amt = client_amount * (transport_pct / 100)

                # Internal cost = Materials + Labour + Misc + O&P + Transport
                internal_cost = base_cost + misc_amt + overhead_amt + transport_amt

                # Cost breakdown section
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = "Cost Analysis:"
                ws[f'A{row}'].font = Font(bold=True, size=10, color="1F4788")
                ws[f'A{row}'].fill = yellow_fill
                row += 1

                # Misc
                ws.cell(row=row, column=2).value = f"Misc ({misc_pct:.1f}%)"
                ws.cell(row=row, column=2).font = normal_font
                ws.cell(row=row, column=6).value = round(misc_amt, 2)
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                row += 1

                # O&P
                ws.cell(row=row, column=2).value = f"O&P ({overhead_pct:.1f}%)"
                ws.cell(row=row, column=2).font = normal_font
                ws.cell(row=row, column=6).value = round(overhead_amt, 2)
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                row += 1

                # Transport (if exists)
                if transport_pct > 0:
                    ws.cell(row=row, column=2).value = f"Transport ({transport_pct:.1f}%)"
                    ws.cell(row=row, column=2).font = normal_font
                    ws.cell(row=row, column=6).value = round(transport_amt, 2)
                    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                    row += 1

                # Total Internal Cost
                ws.cell(row=row, column=5).value = "Total Internal Cost:"
                ws.cell(row=row, column=5).font = Font(bold=True, size=10)
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).value = round(internal_cost, 2)
                ws.cell(row=row, column=6).font = Font(bold=True, size=10)
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                ws.cell(row=row, column=6).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                row += 2

                # Profit Analysis
                # Planned Profit
                ws.cell(row=row, column=5).value = "Planned Profit:"
                ws.cell(row=row, column=5).font = Font(bold=True, size=9, color="10B981")
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).value = round(overhead_amt, 2)
                ws.cell(row=row, column=6).font = Font(bold=True, size=9, color="10B981")
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                row += 1

                # Negotiable Margins (as per BOQDetailsModal)
                negotiable_margin = client_amount - internal_cost
                profit_color = "10B981" if negotiable_margin >= overhead_amt else "EF4444"

                ws.cell(row=row, column=5).value = "Negotiable Margins:"
                ws.cell(row=row, column=5).font = Font(bold=True, size=9, color=profit_color)
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).value = round(negotiable_margin, 2)
                ws.cell(row=row, column=6).font = Font(bold=True, size=9, color=profit_color)
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                row += 3

        row += 2

    # Summary section
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "PROJECT SUMMARY"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Calculate totals (same as PDF)
    items_client_amount = 0
    total_misc = 0
    total_overhead = 0
    total_transport = 0

    for item in items:
        if item.get('has_sub_items') and item.get('sub_items'):
            for sub_item in item.get('sub_items', []):
                qty = sub_item.get('quantity', 0)
                rate = sub_item.get('rate', 0)
                sub_client = qty * rate
                items_client_amount += sub_client

                # Get percentages
                misc_pct = sub_item.get('misc_percentage', item.get('miscellaneous_percentage', item.get('overhead_percentage', 10)))
                overhead_pct = sub_item.get('overhead_profit_percentage', item.get('overhead_profit_percentage', item.get('profit_margin_percentage', 25)))
                transport_pct = sub_item.get('transport_percentage', item.get('transport_percentage', 5))

                # Calculate based on CLIENT AMOUNT
                total_misc += sub_client * (misc_pct / 100)
                total_overhead += sub_client * (overhead_pct / 100)
                total_transport += sub_client * (transport_pct / 100)

    # Extract preliminary amount from boq_json
    preliminary_amount = 0
    if boq_json:
        preliminaries_data = boq_json.get('preliminaries', {})
        cost_details = preliminaries_data.get('cost_details', {})
        preliminary_amount = cost_details.get('amount', 0) or 0

    # Calculate combined client amount (items + preliminary)
    combined_client_amount = items_client_amount + preliminary_amount

    # Get discount from BOQ JSON (same as PDF)
    discount_amount = 0
    discount_percentage = 0

    if boq_json:
        discount_amount = boq_json.get('discount_amount', 0)
        discount_percentage = boq_json.get('discount_percentage', 0)

    # Fallback: try from first item
    if discount_amount == 0 and discount_percentage == 0 and items:
        first_item = items[0]
        discount_percentage = first_item.get('discount_percentage', 0)

    # Calculate discount from combined client amount (items + preliminary) if percentage exists
    if discount_percentage > 0 and discount_amount == 0:
        discount_amount = combined_client_amount * (discount_percentage / 100)

    # Client amount after discount
    client_amount_after_discount = combined_client_amount - discount_amount

    # Internal cost = Materials + Labour + Misc + O&P + Transport
    internal_cost_total = total_material_cost + total_labour_cost + total_misc + total_overhead + total_transport

    # Actual profit = combined client amount - internal
    negotiable_margin_total = combined_client_amount - internal_cost_total

    # Summary data
    summary_items = [
        ("Items Client Amount:", items_client_amount, bold_font),
    ]

    # Add preliminary amount if it exists
    if preliminary_amount > 0:
        summary_items.append(("Preliminary Amount:", preliminary_amount, bold_font))
        summary_items.append(("Combined Client Amount (Excluding VAT):", combined_client_amount, Font(bold=True, size=11)))

    # Add discount if exists
    if discount_amount > 0:
        summary_items.append((f"Discount ({discount_percentage:.1f}%):", discount_amount, Font(bold=True, color="EF4444")))
        summary_items.append(("Client Amount After Discount (Excluding VAT):", client_amount_after_discount, bold_font))

    summary_items.extend([
        ("", 0, normal_font),  # Blank row
        ("Internal Costs:", 0, bold_font),
        ("  - Materials:", total_material_cost, normal_font),
        ("  - Labour:", total_labour_cost, normal_font),
        ("  - Misc:", total_misc, normal_font),
        ("  - O&P:", total_overhead, normal_font),
    ])

    if total_transport > 0:
        summary_items.append(("  - Transport:", total_transport, normal_font))

    summary_items.extend([
        ("Total Internal Cost:", internal_cost_total, Font(bold=True, size=11, color="EF4444")),
        ("", 0, normal_font),  # Blank row
        ("Profit Analysis:", 0, bold_font),
        ("Planned Profit:", total_overhead, Font(bold=True, color="10B981")),
        ("Negotiable Margins (Before Discount):", negotiable_margin_total, Font(bold=True, color="10B981" if negotiable_margin_total >= total_overhead else "EF4444")),
    ])

    # Add actual profit after discount if discount exists
    if discount_amount > 0:
        negotiable_margin_after_discount = client_amount_after_discount - internal_cost_total
        summary_items.append(("Negotiable Margins (After Discount):", negotiable_margin_after_discount, Font(bold=True, color="10B981" if negotiable_margin_after_discount >= total_overhead else "EF4444")))

    summary_items.extend([
        ("", 0, normal_font),  # Blank row
        ("Project Margin:", (negotiable_margin_total / combined_client_amount * 100) if combined_client_amount > 0 else 0, Font(bold=True, size=11)),
    ])

    for label, value, font_style in summary_items:
        if label:
            ws.cell(row=row, column=5).value = label
            ws.cell(row=row, column=5).font = font_style
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='left')

            # Show value if it's non-zero or if it's a section header
            if value != 0 or "Analysis" in label or "Costs" in label:
                ws.cell(row=row, column=6).font = font_style
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')

                # Handle different value types
                if "Margin" in label:
                    # Project Margin is a percentage
                    ws.cell(row=row, column=6).value = round(value, 2)
                    ws.cell(row=row, column=6).number_format = '0.00"%"'
                elif "Discount" in label:
                    # Discount should be negative
                    ws.cell(row=row, column=6).value = -round(value, 2)
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                elif value != 0:
                    # All other values
                    ws.cell(row=row, column=6).value = round(value, 2)
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                else:
                    ws.cell(row=row, column=6).value = ""
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18

    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.read()
