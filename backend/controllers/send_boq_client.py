from flask import request, jsonify, g
from models.boq import BOQ, BOQDetails, BOQHistory, MasterSubItem
from models.project import Project
from config.db import db
from datetime import datetime, date
from config.logging import get_logger
from utils.boq_email_service import BOQEmailService
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from utils.modern_boq_pdf_generator import ModernBOQPDFGenerator
from utils.boq_calculation_helper import calculate_boq_values
from sqlalchemy import text
import os
import time
import threading

log = get_logger()


def _send_boq_background(app, boq_id, client_emails, message, formats, custom_email_body,
                          cover_page, md_signature_image, authorized_signature_image,
                          company_seal_image, estimator_name, estimator_id):
    """Background task: generate files, send emails, update history and notifications."""
    with app.app_context():
        try:
            boq = BOQ.query.filter_by(boq_id=boq_id, is_deleted=False).first()
            if not boq:
                log.error(f"[BG_SEND_BOQ] BOQ {boq_id} not found")
                return

            project = boq.project
            boq_details = BOQDetails.query.filter_by(boq_id=boq_id, is_deleted=False).first()
            if not boq_details:
                log.error(f"[BG_SEND_BOQ] BOQ details not found for {boq_id}")
                return

            boq_json = boq_details.boq_details
            if 'existing_purchase' in boq_json and 'items' in boq_json['existing_purchase']:
                items = boq_json['existing_purchase']['items']
            else:
                items = boq_json.get('items', [])

            total_material_cost, total_labour_cost, items_subtotal, preliminary_amount, grand_total = calculate_boq_values(items, boq_json)

            # Fetch sub_item images
            try:
                sub_item_ids = []
                for item in items:
                    if item.get('has_sub_items'):
                        for sub_item in item.get('sub_items', []):
                            if sub_item.get('sub_item_id'):
                                sub_item_ids.append(sub_item.get('sub_item_id'))
                if sub_item_ids:
                    db_sub_items = MasterSubItem.query.filter(
                        MasterSubItem.sub_item_id.in_(sub_item_ids),
                        MasterSubItem.is_deleted == False
                    ).all()
                    sub_items_map = {si.sub_item_id: si for si in db_sub_items}
                    for item in items:
                        if item.get('has_sub_items'):
                            for sub_item in item.get('sub_items', []):
                                sub_item_id = sub_item.get('sub_item_id')
                                if sub_item_id and sub_item_id in sub_items_map:
                                    db_si = sub_items_map[sub_item_id]
                                    if db_si.sub_item_image:
                                        sub_item['sub_item_image'] = db_si.sub_item_image
                                    if not sub_item.get('description') and db_si.description:
                                        sub_item['description'] = db_si.description
                                    if not sub_item.get('brand') and db_si.brand:
                                        sub_item['brand'] = db_si.brand
                                    if not sub_item.get('size') and db_si.size:
                                        sub_item['size'] = db_si.size
            except Exception as e:
                log.error(f"[BG_SEND_BOQ] Error fetching images: {e}")

            boq_data = {'boq_id': boq.boq_id, 'boq_name': boq.boq_name, 'status': boq.status}
            project_data = {
                'project_name': project.project_name or 'N/A',
                'client': project.client or 'Valued Client',
                'location': project.location or 'N/A'
            }

            # Fetch terms
            selected_terms = []
            try:
                term_ids_result = db.session.execute(
                    text("SELECT term_ids FROM boq_terms_selections WHERE boq_id = :boq_id"),
                    {'boq_id': boq_id}
                ).fetchone()
                term_ids = term_ids_result[0] if term_ids_result and term_ids_result[0] else []
                if term_ids:
                    terms_result = db.session.execute(
                        text("""SELECT terms_text FROM boq_terms
                                WHERE term_id = ANY(:term_ids) AND is_active = TRUE AND is_deleted = FALSE
                                ORDER BY display_order, term_id"""),
                        {'term_ids': term_ids}
                    )
                    for row in terms_result:
                        selected_terms.append({'terms_text': row[0]})
            except Exception as e:
                log.error(f"[BG_SEND_BOQ] Error fetching terms: {e}")

            # Generate files and upload each selected format to Supabase Storage
            excel_file = None
            pdf_file = None
            client_excel_url = None
            client_pdf_url = None

            if 'excel' in formats:
                try:
                    excel_filename = f"BOQ_{project.project_name.replace(' ', '_')}_Client_{date.today().isoformat()}.xlsx"
                    excel_data = generate_client_excel(project, items, total_material_cost, total_labour_cost, grand_total, boq_json, selected_terms)
                    excel_file = (excel_filename, excel_data)
                    client_excel_url = _upload_client_boq_file(excel_data, boq_id, project.project_name, 'xlsx')
                except Exception as e:
                    log.error(f"[BG_SEND_BOQ] Excel generation/upload error: {e}")

            if 'pdf' in formats:
                try:
                    pdf_filename = f"BOQ_{project.project_name.replace(' ', '_')}_Client_{date.today().isoformat()}.pdf"
                    pdf_data = generate_client_pdf(project, items, total_material_cost, total_labour_cost, grand_total, boq_json, selected_terms=selected_terms, include_images=True, cover_page=cover_page, md_signature_image=md_signature_image, authorized_signature_image=authorized_signature_image, company_seal_image=company_seal_image)
                    pdf_file = (pdf_filename, pdf_data)
                    client_pdf_url = _upload_client_boq_file(pdf_data, boq_id, project.project_name, 'pdf')
                except Exception as e:
                    log.error(f"[BG_SEND_BOQ] PDF generation/upload error: {e}")

            # Send emails
            email_service = BOQEmailService()
            successful_sends = []
            failed_sends = []
            for client_email in client_emails:
                try:
                    sent = email_service.send_boq_to_client(
                        boq_data=boq_data,
                        project_data=project_data,
                        client_email=client_email,
                        message=message,
                        total_value=grand_total,
                        item_count=len(items),
                        excel_file=excel_file,
                        pdf_file=pdf_file,
                        custom_email_body=custom_email_body
                    )
                    (successful_sends if sent else failed_sends).append(client_email)
                except Exception as e:
                    failed_sends.append(client_email)
                    log.error(f"[BG_SEND_BOQ] Error sending to {client_email}: {e}")

            log.info(f"[BG_SEND_BOQ] Sent {len(successful_sends)}/{len(client_emails)} for BOQ {boq_id}")

            # Update history
            try:
                existing_history = BOQHistory.query.filter_by(boq_id=boq_id).order_by(BOQHistory.action_date.desc()).first()
                if existing_history:
                    if isinstance(existing_history.action, list):
                        current_actions = existing_history.action
                    elif isinstance(existing_history.action, dict):
                        current_actions = [existing_history.action]
                    else:
                        current_actions = []
                else:
                    current_actions = []

                current_actions.append({
                    "role": "estimator",
                    "type": "sent_to_client",
                    "sender": "estimator",
                    "receiver": "client",
                    "status": "Sent_for_Confirmation",
                    "boq_name": boq.boq_name,
                    "comments": message or "BOQ sent to client for confirmation",
                    "timestamp": datetime.utcnow().isoformat(),
                    "sender_name": estimator_name,
                    "sender_user_id": estimator_id,
                    "total_value": grand_total,
                    "item_count": len(items),
                    "project_name": project.project_name,
                    "client_name": project.client,
                    "client_emails": successful_sends,
                    "failed_emails": failed_sends if failed_sends else None
                })

                recipients_str = ', '.join(successful_sends)
                if existing_history:
                    existing_history.action = current_actions
                    existing_history.action_date = datetime.utcnow()
                    existing_history.comment = f"BOQ sent to {len(successful_sends)} client(s): {recipients_str}"
                else:
                    db.session.add(BOQHistory(
                        boq_id=boq_id,
                        action=current_actions,
                        action_date=datetime.utcnow(),
                        comment=f"BOQ sent to {len(successful_sends)} client(s): {recipients_str}"
                    ))
                db.session.commit()
            except Exception as e:
                log.error(f"[BG_SEND_BOQ] Error updating history: {e}")

            # Send notification to TD
            try:
                from utils.comprehensive_notification_service import notification_service
                from models.user import User, Role
                td_role = Role.query.filter_by(role='technicalDirector').first()
                if td_role:
                    td_users = User.query.filter_by(role_id=td_role.role_id, is_deleted=False, is_active=True).all()
                    td_user_ids = [td.user_id for td in td_users]
                    if td_user_ids:
                        notification_service.notify_boq_sent_to_client(
                            boq_id=boq_id,
                            project_name=project.project_name,
                            estimator_id=estimator_id,
                            estimator_name=estimator_name,
                            td_user_ids=td_user_ids,
                            client_email=', '.join(successful_sends)
                        )
            except Exception as e:
                log.error(f"[BG_SEND_BOQ] Notification error: {e}")

        except Exception as e:
            import traceback
            log.error(f"[BG_SEND_BOQ] Background task failed: {e}")
            log.error(traceback.format_exc())


def send_boq_to_client():
    """
    Send BOQ to client — returns immediately, heavy work runs in background thread.
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400

        boq_id = data.get('boq_id')
        client_emails_raw = data.get('client_email') or data.get('client_emails')
        message = data.get('message', 'Please review the attached BOQ for your project.')
        formats = data.get('formats', ['excel', 'pdf'])
        custom_email_body = data.get('custom_email_body')
        cover_page = data.get('cover_page')
        include_signature = data.get('include_signature', False)

        md_signature_image = None
        authorized_signature_image = None
        company_seal_image = None
        if include_signature:
            from controllers.settings_controller import get_signatures_for_pdf
            signatures = get_signatures_for_pdf()
            md_signature_image = signatures.get('md_signature')
            authorized_signature_image = signatures.get('authorized_signature')
            company_seal_image = signatures.get('company_seal')

        if not boq_id:
            return jsonify({"success": False, "error": "boq_id is required"}), 400

        boq = BOQ.query.filter_by(boq_id=boq_id, is_deleted=False).first()
        if not boq:
            return jsonify({"success": False, "error": "BOQ not found"}), 404

        if boq.status not in ["Approved", "Revision_Approved", "Internal_Revision_Approved"]:
            return jsonify({
                "success": False,
                "error": f"BOQ must be approved before sending to client. Current status: {boq.status}"
            }), 400

        project = boq.project
        if not project:
            return jsonify({"success": False, "error": "Project not found for this BOQ"}), 404

        # Use project.client_email as fallback
        if not client_emails_raw and project.client_email:
            client_emails_raw = project.client_email

        if not client_emails_raw:
            return jsonify({
                "success": False,
                "error": "No client email provided. Please enter a client email or add one to the project."
            }), 400

        # Parse emails
        if isinstance(client_emails_raw, list):
            client_emails = [e.strip() for e in client_emails_raw if e.strip()]
        else:
            client_emails = [e.strip() for e in client_emails_raw.replace(';', ',').split(',') if e.strip()]

        if not client_emails:
            return jsonify({"success": False, "error": "At least one valid email address is required"}), 400

        import re
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        invalid_emails = [e for e in client_emails if not email_pattern.match(e)]
        if invalid_emails:
            return jsonify({"success": False, "error": f"Invalid email format: {', '.join(invalid_emails)}"}), 400

        current_user = getattr(g, 'user', None)
        estimator_name = current_user.get('full_name', 'Estimator') if current_user else 'Estimator'
        estimator_id = current_user.get('user_id') if current_user else None

        # Immediately update BOQ status and save client email — commit before returning
        boq.email_sent = True
        boq.client_status = False
        boq.status = "Sent_for_Confirmation"
        boq.last_modified_by = estimator_name
        boq.last_modified_at = datetime.utcnow()
        project.client_email = ', '.join(client_emails)
        db.session.commit()

        # Fire background thread for PDF generation, email sending, history, notifications
        from flask import current_app
        app = current_app._get_current_object()
        thread = threading.Thread(
            target=_send_boq_background,
            args=(app, boq_id, client_emails, message, formats, custom_email_body,
                  cover_page, md_signature_image, authorized_signature_image,
                  company_seal_image, estimator_name, estimator_id),
            daemon=True
        )
        thread.start()

        return jsonify({
            "success": True,
            "message": f"BOQ is being sent to {len(client_emails)} recipient(s). The email will arrive shortly.",
            "boq_id": boq_id,
            "status": "Sent_for_Confirmation",
            "client_emails": client_emails,
        }), 200

    except Exception as e:
        import traceback
        log.error(f"Error sending BOQ to client: {str(e)}")
        log.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500


def _upload_client_boq_file(file_bytes, boq_id, project_name, file_ext):
    """Upload a client BOQ file (pdf or xlsx) to Supabase Storage boq_file/client_boq/ and return public URL."""
    try:
        from supabase import create_client as create_supabase_client

        SUPABASE_BUCKET = "boq_file"
        environment = os.environ.get('ENVIRONMENT', 'production')

        if environment == 'development':
            supabase_url = os.environ.get('DEV_SUPABASE_URL')
            upload_key = os.environ.get('DEV_SUPABASE_KEY')
            anon_key = os.environ.get('DEV_SUPABASE_ANON_KEY')
        else:
            supabase_url = os.environ.get('SUPABASE_URL')
            upload_key = os.environ.get('SUPABASE_KEY')
            anon_key = os.environ.get('SUPABASE_ANON_KEY')

        if not supabase_url or not upload_key:
            log.error(f"[SEND_BOQ] Upload: Missing Supabase credentials")
            return None

        content_types = {
            'pdf': 'application/pdf',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }

        upload_client = create_supabase_client(supabase_url, upload_key)
        timestamp = int(time.time())
        safe_project_name = project_name.replace(' ', '_').replace('/', '-')[:40]
        filename = f"BOQ_{safe_project_name}_{timestamp}.{file_ext}"
        file_path = f"client_boq/boq_{boq_id}/{filename}"

        upload_client.storage.from_(SUPABASE_BUCKET).upload(
            file_path,
            file_bytes,
            {
                "content-type": content_types.get(file_ext, 'application/octet-stream'),
                "content-disposition": f'inline; filename="{filename}"',
                "x-upsert": "true",
            }
        )

        anon_client = create_supabase_client(supabase_url, anon_key or upload_key)
        file_url = anon_client.storage.from_(SUPABASE_BUCKET).get_public_url(file_path)

        log.info(f"[SEND_BOQ] Uploaded {file_ext} to: {file_path}")
        return file_url

    except Exception as e:
        log.error(f"[SEND_BOQ] Upload error ({file_ext}): {e}")
        import traceback
        log.error(traceback.format_exc())
        return None


# Keep backward-compatible alias
def _upload_client_boq_pdf(pdf_bytes, boq_id, project_name):
    return _upload_client_boq_file(pdf_bytes, boq_id, project_name, 'pdf')


def generate_client_excel(project, items, total_material_cost, total_labour_cost, client_base_cost, boq_json=None, selected_terms=None):
    """
    Generate Client Excel file - MODERN PROFESSIONAL FORMAT
    Shows ONLY items and sub-items (NO raw materials/labour details)
    Overhead & Profit already included in selling prices

    Args:
        selected_terms: List of selected terms from database. Each dict should have {'terms_text': '...'}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    if boq_json is None:
        boq_json = {}

    # Define styles
    header_font = Font(bold=True, size=16, color="1F4788")
    sub_header_font = Font(bold=True, size=12, color="1F4788")
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    total_font = Font(bold=True, size=12, color="10B981")

    # Header fills
    blue_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    grey_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title - Updated for image column
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "QUOTATION"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "Bill of Quantities"
    ws[f'A{row}'].font = Font(size=11, italic=True, color="6B7280")
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Project Information Box
    ws[f'A{row}'] = "Project Information"
    ws[f'A{row}'].font = sub_header_font
    row += 1

    info_data = [
        ["Project Name:", project.project_name],
        ["Client:", project.client],
        ["Location:", project.location],
        ["Date:", date.today().strftime('%d %B %Y')]
    ]

    for info_row in info_data:
        ws[f'A{row}'] = info_row[0]
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = grey_fill
        ws[f'B{row}'] = info_row[1]
        ws[f'B{row}'].font = normal_font
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    row += 2

    # Scope of Work Header
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "SCOPE OF WORK"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Process each item
    for idx, item in enumerate(items, 1):
        # Item Header
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"{idx}. {item.get('item_name', 'N/A')}"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="1F4788")
        ws[f'A{row}'].fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        row += 1

        if item.get('description'):
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = item['description']
            ws[f'A{row}'].font = Font(italic=True, size=9, color="6B7280")
            row += 1

        row += 1

        # Check if item has sub-items
        has_sub_items = item.get('has_sub_items', False)
        sub_items = item.get('sub_items', [])

        # Initialize item total (used later in line 438)
        item_total_calculated = 0

        if has_sub_items and sub_items:
            # Sub-items table header - Added Image column
            headers = ['Sub-Item Description', 'Scope / Size', 'Image', 'Qty', 'Unit', 'Rate (AED)', 'Amount (AED)']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = table_header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            row += 1

            # Calculate item total from sub-items (clean calculation)

            # Sub-items data (CLEAN CLIENT VIEW - just qty × rate)
            for sub_item in sub_items:
                sub_item_name = sub_item.get('sub_item_name', 'N/A')
                scope = sub_item.get('scope', '')
                size = sub_item.get('size', '')
                location = sub_item.get('location', '')
                brand = sub_item.get('brand', '')
                spec = sub_item.get('description', '')

                # Build scope/size display
                scope_parts = []
                if scope:
                    scope_parts.append(scope)
                if size:
                    scope_parts.append(f"Size: {size}")
                if location:
                    scope_parts.append(f"Loc: {location}")
                if brand:
                    scope_parts.append(f"Brand: {brand}")
                if spec:
                    scope_parts.append(f"Spec: {spec}")
                scope_size = " | ".join(scope_parts) if scope_parts else '-'

                # CLIENT VIEW: Simple qty × rate (same as PDF)
                quantity = sub_item.get('quantity', 0)
                unit = sub_item.get('unit', 'nos')
                rate = sub_item.get('rate', 0)
                sub_item_total = quantity * rate

                # Accumulate for item total
                item_total_calculated += sub_item_total

                # Get all images from sub_item_image JSONB array
                image_text = ''
                sub_item_images = sub_item.get('sub_item_image', [])
                if sub_item_images and isinstance(sub_item_images, list) and len(sub_item_images) > 0:
                    # Count total images and create link text
                    image_count = len(sub_item_images)
                    if image_count == 1:
                        first_image = sub_item_images[0]
                        if isinstance(first_image, dict):
                            image_url = first_image.get('url', '')
                            if image_url:
                                image_text = 'View Image'
                    else:
                        # Multiple images - just use the first one for the link
                        first_image = sub_item_images[0]
                        if isinstance(first_image, dict):
                            image_url = first_image.get('url', '')
                            if image_url:
                                image_text = f'View Images ({image_count})'

                # Write row (updated with image column)
                ws.cell(row=row, column=1).value = sub_item_name
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=row, column=2).value = scope_size
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

                # Image column (column 3)
                image_cell = ws.cell(row=row, column=3)
                if image_text and image_url:
                    image_cell.value = image_text
                    image_cell.hyperlink = image_url
                    image_cell.font = Font(color="0563C1", underline="single")
                image_cell.alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(row=row, column=4).value = round(quantity, 2)
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=5).value = unit
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=6).value = round(rate, 2)
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                ws.cell(row=row, column=7).value = round(sub_item_total, 2)
                ws.cell(row=row, column=7).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=7).number_format = '#,##0.00'

                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).font = normal_font

                row += 1

        else:
            # Old format: No sub-items
            item_qty = item.get('quantity', 0)
            item_unit = item.get('unit', 'nos')
            item_rate = item.get('rate', 0)
            item_total = item.get('selling_price', 0)

            ws[f'A{row}'] = f"Quantity: {item_qty} {item_unit}"
            ws[f'D{row}'] = f"Rate: AED {item_rate:,.2f}/{item_unit}"
            row += 1

        # Item Total - Updated to merge to column F (added image column)
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Item Total:"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="10B981")
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        # Use calculated total if has sub-items, otherwise use old field
        item_total_value = item_total_calculated if (item.get('has_sub_items', False) and item.get('sub_items', [])) else item.get('selling_price', 0)
        ws[f'G{row}'] = round(item_total_value, 2)
        ws[f'G{row}'].font = Font(bold=True, size=11, color="10B981")
        ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        row += 3

    # Cost Summary
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "COST SUMMARY"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Calculate items subtotal from sub-items (qty × rate) - same as PDF
    items_subtotal = 0
    for item in items:
        has_sub_items = item.get('has_sub_items', False)
        sub_items = item.get('sub_items', [])

        if has_sub_items and sub_items:
            for sub_item in sub_items:
                qty = sub_item.get('quantity', 0)
                rate = sub_item.get('rate', 0)
                items_subtotal += qty * rate
        else:
            items_subtotal += item.get('selling_price', 0)

    # Extract preliminary amount from boq_json
    preliminary_amount = 0
    if boq_json:
        preliminaries_data = boq_json.get('preliminaries', {})
        cost_details = preliminaries_data.get('cost_details', {})
        preliminary_amount = cost_details.get('amount', 0) or 0

    # Calculate combined subtotal (items + preliminary)
    combined_subtotal = items_subtotal + preliminary_amount

    # Get discount from BOQ JSON (same as PDF fix)
    discount_amount = 0
    discount_percentage = 0

    if boq_json:
        discount_amount = boq_json.get('discount_amount', 0)
        discount_percentage = boq_json.get('discount_percentage', 0)

    # Fallback: try from first item
    if discount_amount == 0 and discount_percentage == 0 and items:
        first_item = items[0]
        discount_percentage = first_item.get('discount_percentage', 0)

    # Calculate discount from combined subtotal (items + preliminary) if percentage exists
    if discount_percentage > 0 and discount_amount == 0:
        discount_amount = combined_subtotal * (discount_percentage / 100)

    total_discount = discount_amount
    subtotal_after_discount = combined_subtotal - total_discount
    total_vat = 0  # VAT not used
    grand_total_with_vat = subtotal_after_discount + total_vat

    # Items Subtotal - Updated for image column
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "Items Subtotal:"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'G{row}'] = round(items_subtotal, 2)
    ws[f'G{row}'].font = bold_font
    ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'G{row}'].number_format = '#,##0.00'
    row += 1

    # Preliminary Amount (if exists)
    if preliminary_amount > 0:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Preliminary Amount:"
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'] = round(preliminary_amount, 2)
        ws[f'G{row}'].font = bold_font
        ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'].number_format = '#,##0.00'
        row += 1

        # Combined Subtotal
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Combined Subtotal:"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'] = round(combined_subtotal, 2)
        ws[f'G{row}'].font = Font(bold=True, size=11)
        ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'].number_format = '#,##0.00'
        row += 1

    # Discount (only if exists)
    if discount_amount > 0:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Discount ({discount_percentage:.1f}%):"
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'] = -round(discount_amount, 2)
        ws[f'G{row}'].font = Font(bold=True, color="EF4444")
        ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'].number_format = '#,##0.00'
        row += 1

        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "After Discount:"
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'] = round(subtotal_after_discount, 2)
        ws[f'G{row}'].font = bold_font
        ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'G{row}'].number_format = '#,##0.00'
        row += 1

    # VAT row removed (not used)
    row += 1

    # Grand Total
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "TOTAL PROJECT VALUE (Excluding VAT):"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = green_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'G{row}'] = round(grand_total_with_vat, 2)
    ws[f'G{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'G{row}'].fill = green_fill
    ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'G{row}'].number_format = '#,##0.00'
    row += 2

    # Preliminaries
    preliminaries = boq_json.get('preliminaries', {})
    prelim_items = preliminaries.get('items', [])
    prelim_notes = preliminaries.get('notes', '')

    if prelim_items or prelim_notes:
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "PRELIMINARIES & APPROVAL WORKS"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="643CCA")
        row += 1

        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "Selected conditions and terms"
        ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")
        row += 2

        for prelim_item in prelim_items:
            # Show all preliminaries (both selected and unselected)
            if isinstance(prelim_item, dict):
                desc = prelim_item.get('description', prelim_item.get('name', prelim_item.get('text', '')))
                is_selected = prelim_item.get('is_selected', prelim_item.get('selected', prelim_item.get('checked', False)))
                # Use checkmark for selected, circle for unselected
                prefix = "✓ " if is_selected else "○ "
            else:
                desc = str(prelim_item)
                prefix = "✓ "

            if desc:  # Only add if text exists
                ws.merge_cells(f'A{row}:G{row}')
                ws[f'A{row}'] = f"{prefix}{desc}"
                ws[f'A{row}'].font = normal_font
                row += 1

        if prelim_notes:
            row += 1
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "Additional Notes:"
            ws[f'A{row}'].font = bold_font
            row += 1
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = prelim_notes
            ws[f'A{row}'].font = Font(italic=True, size=9)
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)

    # Terms & Conditions Section
    if selected_terms and len(selected_terms) > 0:
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "TERMS & CONDITIONS"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="643CCA")
        row += 1

        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "Selected terms and conditions for this quotation"
        ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")
        row += 1

        for idx, term in enumerate(selected_terms, 1):
            term_text = term.get('terms_text', '').strip()
            if term_text:  # Only add if text exists
                ws.merge_cells(f'A{row}:G{row}')
                ws[f'A{row}'] = f"• {term_text}"
                ws[f'A{row}'].font = normal_font
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1

    # Column widths - Updated for image column
    ws.column_dimensions['A'].width = 28  # Sub-Item Description
    ws.column_dimensions['B'].width = 22  # Scope / Size
    ws.column_dimensions['C'].width = 12  # Image (new)
    ws.column_dimensions['D'].width = 8   # Qty
    ws.column_dimensions['E'].width = 8   # Unit
    ws.column_dimensions['F'].width = 15  # Rate
    ws.column_dimensions['G'].width = 18  # Amount

    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.read()


def generate_client_pdf(project, items, total_material_cost, total_labour_cost, grand_total, boq_json=None, selected_terms=None, include_images=True, cover_page=None, md_signature_image=None, authorized_signature_image=None, company_seal_image=None):
    """
    Generate Client PDF - MODERN PROFESSIONAL CORPORATE FORMAT
    Uses unified ModernBOQPDFGenerator

    Args:
        selected_terms: List of selected terms from database. Each dict should have {'terms_text': '...'}
        include_images: If True, include images (slower). Default False for email speed.
        cover_page: Optional dict with cover page data for quotation letter
        md_signature_image: MD signature for cover page (base64)
        authorized_signature_image: Authorized signature for quotation section (base64)
        company_seal_image: Company seal/stamp image (base64)
    """
    if boq_json is None:
        boq_json = {}

    generator = ModernBOQPDFGenerator()
    return generator.generate_client_pdf(project, items, total_material_cost, total_labour_cost, grand_total, boq_json, terms_text=None, selected_terms=selected_terms, include_images=include_images, cover_page=cover_page, md_signature_image=md_signature_image, authorized_signature_image=authorized_signature_image, company_seal_image=company_seal_image)
