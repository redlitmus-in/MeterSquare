from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import hashlib
import uuid
import pandas as pd
import pdfplumber
import PyPDF2
import json
import re
from config.db import db
from config.logging import get_logger
from models.boq import BOQ, BOQDetails, MasterItem, MasterMaterial, MasterLabour
from models.project import Project
from utils.pdf_extractor import PDFExtractor, extract_boq_from_pdf
from supabase import create_client, Client
from dotenv import load_dotenv

load_dotenv()
log = get_logger()

estimator_boq_bp = Blueprint('estimator_boq', __name__)

# Supabase configuration based on ENVIRONMENT variable
environment = os.environ.get('ENVIRONMENT', 'production')
if environment == 'development':
    SUPABASE_URL = os.getenv('DEV_SUPABASE_URL')
    SUPABASE_KEY = os.getenv('DEV_SUPABASE_ANON_KEY')
else:
    SUPABASE_URL = os.getenv('SUPABASE_URL')
    SUPABASE_KEY = os.getenv('SUPABASE_ANON_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

# File upload configuration
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

# Remove LLM dependencies - using rule-based extraction for better reliability

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_hash(file_content):
    """Generate SHA256 hash of file content"""
    return hashlib.sha256(file_content).hexdigest()

def extract_text_from_pdf(file_path):
    """Extract text from PDF file using enhanced PDF extractor"""
    try:
        # Use the enhanced PDF extractor
        extractor = PDFExtractor(file_path)
        result = extractor.extract()

        if result["success"]:
            text = ""
            tables = []

            # Process extracted data
            for page in result.get("pages", []):
                text += f"\n--- Page {page['page_number']} ---\n{page.get('text', '')}"
                tables.extend(page.get('tables', []))

            return text, tables, result.get("boq_data", {})
        else:
            # Fallback to basic extraction
            return extract_text_from_pdf_fallback(file_path)
    except Exception as e:
        log.error(f"Enhanced PDF extraction failed: {e}")
        return extract_text_from_pdf_fallback(file_path)

def extract_text_from_pdf_fallback(file_path):
    """Fallback PDF extraction method"""
    text = ""
    tables = []
    boq_data = {}

    try:
        # Method 1: pdfplumber for text and tables
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                # Extract text
                page_text = page.extract_text()
                if page_text:
                    text += f"\n--- Page {i+1} ---\n{page_text}"

                # Extract tables with better settings
                page_tables = page.extract_tables(
                    table_settings={
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "snap_tolerance": 3,
                        "join_tolerance": 3,
                        "text_tolerance": 3
                    }
                )

                # If no tables found with lines, try text-based extraction
                if not page_tables:
                    page_tables = page.extract_tables(
                        table_settings={
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text"
                        }
                    )

                for table in page_tables:
                    if table:
                        # Clean up the table data
                        cleaned_table = []
                        for row in table:
                            cleaned_row = []
                            for cell in row:
                                if cell is None:
                                    cleaned_row.append('')
                                else:
                                    cleaned_cell = str(cell).strip()
                                    cleaned_cell = ' '.join(cleaned_cell.split())
                                    cleaned_row.append(cleaned_cell)
                            cleaned_table.append(cleaned_row)

                        tables.append({
                            'page': i + 1,
                            'data': cleaned_table
                        })
    except Exception as e:
        log.error(f"pdfplumber extraction failed: {e}")
        # Fallback: PyPDF2
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for i, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text:
                        text += f"\n--- Page {i+1} ---\n{page_text}"
        except Exception as e2:
            log.error(f"PyPDF2 extraction failed: {e2}")

    return text, tables, boq_data

def extract_text_from_excel(file_path):
    """Extract text and data from Excel file"""
    text = ""
    tables = []

    try:
        # Read Excel file with openpyxl engine (for .xlsx files)
        # For .xls files, we'll need xlrd but xlrd doesn't work with Python 3.12+
        file_ext = file_path.rsplit('.', 1)[1].lower()

        if file_ext == 'xls':
            return text, tables
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            log.error(f"Error loading workbook: {e}")
            # Fallback: try reading with pandas directly
            sheet_names = [0]  # Read first sheet by index

        for sheet_name in sheet_names:
            # Read with header detection - directly specify openpyxl engine
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            # Find the header row by looking for key column names
            header_row_idx = None
            for idx, row in df.iterrows():
                row_str = ' '.join([str(x).lower() for x in row if pd.notna(x)])
                if any(keyword in row_str for keyword in ['work type', 'description', 'qty', 'unit', 'rate', 'amount', 'item', 'sub item']):
                    header_row_idx = idx
                    break

            if header_row_idx is not None:
                # Use found header row
                headers = df.iloc[header_row_idx].tolist()
                data_rows = df.iloc[header_row_idx + 1:].values.tolist()
            else:
                # No header found, use first row
                headers = df.iloc[0].tolist()
                data_rows = df.iloc[1:].values.tolist()

            # Clean headers and data
            cleaned_headers = []
            for h in headers:
                if pd.notna(h):
                    h_str = str(h).strip()
                    # Handle nan strings from Excel
                    if h_str.lower() != 'nan':
                        cleaned_headers.append(h_str)
                    else:
                        cleaned_headers.append('')
                else:
                    cleaned_headers.append('')

            cleaned_data = []
            for row in data_rows:
                cleaned_row = []
                for cell in row:
                    if pd.notna(cell):
                        cell_str = str(cell).strip()
                        # Handle nan strings from Excel
                        if cell_str.lower() != 'nan':
                            cleaned_row.append(cell_str)
                        else:
                            cleaned_row.append('')
                    else:
                        cleaned_row.append('')

                # Skip completely empty rows
                if any(cell for cell in cleaned_row):
                    cleaned_data.append(cleaned_row)

            # Convert to text
            text += f"\n--- Sheet: {sheet_name} ---\n"
            text += f"Headers: {cleaned_headers}\n"
            for row in cleaned_data[:10]:  # Sample first 10 rows for text
                text += f"{row}\n"

            # Store as table with proper structure
            table_data = [cleaned_headers] + cleaned_data
            tables.append({
                'sheet': sheet_name,
                'data': table_data  # This matches the format expected by process_boq_table
            })
    except Exception as e:
        log.error(f"Excel extraction failed: {e}")
    return text, tables

def extract_boq_data(text, tables, boq_data=None):
    """Extract BOQ data using enhanced rule-based extraction"""
    # If we already have BOQ data from PDF extractor, use it
    if boq_data and boq_data.get('sections'):
        return format_extracted_boq_data(boq_data)

    # Otherwise, use rule-based extraction
    return extract_with_rules(text, tables)

def format_extracted_boq_data(boq_data):
    """Format BOQ data from PDF extractor to our standard format"""
    extracted = {
        'project_details': boq_data.get('project_info', {}),
        'boq_items': []
    }

    # Convert sections to flat list of items
    for section in boq_data.get('sections', []):
        work_type = section.get('name', 'General')
        for item in section.get('items', []):
            extracted['boq_items'].append({
                'work_type': work_type,
                'item_name': item.get('description', ''),
                'description': item.get('description', ''),
                'quantity': float(item.get('quantity', 0)),
                'unit': item.get('unit', 'nos'),
                'rate': float(item.get('rate', 0)),
                'amount': float(item.get('amount', 0))
            })

    return extracted

def extract_with_rules(text, tables):
    """Rule-based extraction as fallback when LLM is not available"""
    extracted = {
        'project_details': {},
        'boq_items': []
    }

    # Extract project information using regex patterns
    patterns = {
        'project_name': r'Project\s*[Nn]ame\s*[:：]\s*([^\n]+)',
        'client': r'Client\s*[:：]\s*([^\n]+)',
        'location': r'Location\s*[:：]\s*([^\n]+)',
        'work_type': r'Work\s*[Tt]ype\s*[:：]\s*([^\n]+)',
        'area': r'Area\s*[:：]\s*([0-9]+\s*[A-Za-z]+)'
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted['project_details'][key] = match.group(1).strip()

    # Process tables to extract BOQ items
    for idx, table_data in enumerate(tables):
        if 'data' in table_data:
            items = process_boq_table(table_data['data'])
            extracted['boq_items'].extend(items)
        else:
            log.debug(f"Table #{idx + 1} has no 'data' key, skipping")

    return extracted

def process_boq_table(table_data):
    """Process a table to extract BOQ items according to PDF format"""
    items = []

    if not table_data or len(table_data) < 2:
        return items

    # Clean and normalize headers
    headers = []
    for h in table_data[0]:
        if h is None:
            headers.append('')
        else:
            # Clean the header text
            header = str(h).lower().strip()
            # Remove extra whitespaces and special chars
            header = ' '.join(header.split())
            # Remove parentheses content for easier matching
            header = header.replace('(aed)', '').replace('(', '').replace(')', '').strip()
            headers.append(header)
    # If no valid headers found, skip this table
    if not any(headers):
        return items

    # Find relevant column indices - matching your Excel format exactly
    col_map = {}
    labour_role_idx = None  # Track labour role column to identify the next Amount column
    amount_columns = []  # Track all amount columns to distinguish between material and labour amounts

    for i, h in enumerate(headers):
        h_clean = str(h).strip()
        h_lower = h_clean.lower()

        # Remove common patterns to simplify matching
        h_normalized = h_lower.replace('(aed)', '').replace('aed', '').replace('(', '').replace(')', '').strip()
        # Work type
        if 'work' in h_normalized and 'type' in h_normalized:
            col_map['work_type'] = i
        # Item column (exact match to avoid confusion with "Sub Item")
        elif h_normalized == 'item':
            col_map['item'] = i
        # Sub item column
        elif 'sub' in h_normalized and 'item' in h_normalized:
            col_map['sub_item'] = i
        # Description
        elif 'description' in h_normalized:
            col_map['description'] = i
        # QTY
        elif h_normalized in ['qty', 'quantity']:
            col_map['qty'] = i
        # Unit
        elif h_normalized == 'unit':
            col_map['unit'] = i
        # Rate - Material rate (before labour columns, first rate column)
        elif 'rate' in h_normalized and 'per' not in h_normalized and 'hour' not in h_normalized and 'rate' not in col_map:
            col_map['rate'] = i
        # Amount - Material amount (appears before Labour Role)
        elif 'amount' in h_normalized and labour_role_idx is None and 'amount' not in col_map:
            col_map['amount'] = i
            amount_columns.append(('material', i))
        # Labour Role
        elif 'labour' in h_normalized and 'role' in h_normalized:
            col_map['labour_role'] = i
            labour_role_idx = i
        # Working Hours
        elif 'working' in h_normalized and 'hours' in h_normalized:
            col_map['working_hours'] = i
        # Profit margin percentage
        elif 'profit' in h_normalized and 'margin' in h_normalized:
            col_map['profit_margin_percentage'] = i
        # Overhead percentage (comes after profit in your format)
        elif 'overhead' in h_normalized or 'over' in h_normalized and 'head' in h_normalized:
            col_map['overhead_percentage'] = i
        # Amount column after Labour Role = Labour Amount
        elif 'amount' in h_normalized and labour_role_idx is not None and i > labour_role_idx:
            if 'labour_amount' not in col_map:
                col_map['labour_amount'] = i
                amount_columns.append(('labour', i))

    # Second pass: If labour_amount not found but we have labour_role, find next "Amount" column
    if labour_role_idx is not None and 'labour_amount' not in col_map:
        for i, h in enumerate(headers):
            h_lower = h.lower()
            # Look for "Amount" column that appears after "Labour Role"
            if i > labour_role_idx and h_lower == 'amount':
                col_map['labour_amount'] = i
                break

    for field_name, col_idx in col_map.items():
        actual_header = headers[col_idx] if col_idx < len(headers) else 'N/A'
    # Extract items from rows
    item_counter = 1
    current_work_type = "General"  # Default work type
    current_main_item = None

    for row in table_data[1:]:
        if not row or all(not cell for cell in row):
            continue
        # Extract all fields into item dict
        item = {}
        for field, idx in col_map.items():
            if idx is not None and idx < len(row):
                value = row[idx]
                if value and str(value).strip():
                    # Clean and convert values
                    if field in ['qty', 'rate', 'amount', 'labour_amount', 'rate_per_hour', 'working_hours',
                                 'profit_margin_percentage', 'overhead_percentage']:
                        # Extract numeric value
                        value = str(value).replace(',', '').replace('AED', '').replace('%', '').strip()
                        try:
                            item[field] = float(value)
                        except:
                            item[field] = 0.0
                    else:
                        item[field] = str(value).strip()
        # Get main values
        item_value = item.get('item', '')
        sub_item_value = item.get('sub_item', '')
        description_value = item.get('description', '')
        work_type_value = item.get('work_type', '')
        # Update work type if present in this row
        if work_type_value and str(work_type_value).strip():
            current_work_type = str(work_type_value).strip()
        # In your format: Item column = main category, Sub Item column = actual MATERIAL
        if sub_item_value and str(sub_item_value).strip():
            # Sub-item has value - this is the MATERIAL (sub item = material)
            item['item_name'] = str(sub_item_value).strip()
            # Use Item column as parent category
            if item_value and str(item_value).strip():
                current_main_item = str(item_value).strip()
                item['parent_item'] = current_main_item
            else:
                item['parent_item'] = current_main_item if current_main_item else 'General'
            item['is_sub_item'] = True
            item['is_material'] = True  # Sub Item = Material (as per user requirement)
        elif item_value and str(item_value).strip():
            # Only Item column has value - could be a category header or standalone item
            # Check if it has financial data (qty, rate, amount)
            has_financial_data = (
                item.get('qty', 0) > 0 or
                item.get('rate', 0) > 0 or
                item.get('amount', 0) > 0
            )
            if has_financial_data:
                # It's a standalone item with data
                item['item_name'] = str(item_value).strip()
                item['parent_item'] = current_work_type
                item['is_sub_item'] = False
            else:
                # This is a category header row - track it but don't add as item
                current_main_item = str(item_value).strip()
                continue  # Skip to next row
        elif description_value and str(description_value).strip():
            # Use description as item name
            item['item_name'] = str(description_value).strip()
            item['parent_item'] = current_main_item if current_main_item else current_work_type
            item['is_sub_item'] = False
        else:
            # No meaningful data in this row
            continue

        # Add description if not already set
        if 'description' not in item or not item['description']:
            item['description'] = item.get('item_name', '')

        # Set item properties
        item['item_number'] = str(item_counter)
        item['work_type'] = current_work_type

        # Store labour information properly
        if 'labour_role' in item and item.get('labour_role'):
            working_hours = item.get('working_hours', 0.0)
            labour_amount = item.get('labour_amount', 0.0)
            rate_per_hour = 0.0

            # Calculate rate_per_hour from labour_amount and working_hours
            if labour_amount > 0 and working_hours > 0:
                rate_per_hour = labour_amount / working_hours
            item['labour'] = {
                'role': item.get('labour_role', ''),
                'amount': labour_amount,
                'working_hours': working_hours,
                'rate_per_hour': rate_per_hour
            }

        # Only add if we have valid data (item name and at least some financial data)
        if item.get('item_name') and item.get('item_name').strip():
            # Check if we have any financial data (material or labour)
            has_material_data = (
                item.get('rate', 0) > 0 or
                item.get('amount', 0) > 0 or
                item.get('qty', 0) > 0
            )
            has_labour_data = 'labour' in item and item.get('labour', {}).get('amount', 0) > 0
            has_financial_data = has_material_data or has_labour_data

            if has_financial_data:
                has_labour = 'labour' in item and item['labour']
                if has_labour:
                    log.debug(f"Labour details: {item['labour']}")
                items.append(item)
                item_counter += 1
            else:
                log.debug(f"Skipping item '{item.get('item_name')}' - no financial data")

    return items

def truncate_and_clean_name(name, max_length=100):
    """Truncate and clean a name to fit database constraints"""
    if not name:
        return "Untitled BOQ"

    # Remove newlines and excessive whitespace
    name = ' '.join(name.split())

    # Truncate if too long
    if len(name) > max_length:
        name = name[:max_length-3] + "..."

    return name

def validate_and_clean_data(extracted_data):
    """Validate and clean extracted BOQ data"""
    # Ensure we have required fields
    if 'project_details' not in extracted_data:
        extracted_data['project_details'] = {}

    if 'boq_items' not in extracted_data:
        extracted_data['boq_items'] = []

    # Clean project name if it exists
    if 'project_name' in extracted_data['project_details']:
        original_name = extracted_data['project_details']['project_name']
        extracted_data['project_details']['project_name'] = truncate_and_clean_name(original_name, max_length=200)

        # Store original name separately if truncated
        if len(original_name) > 200:
            extracted_data['project_details']['original_project_name'] = original_name

    # Clean and validate items
    cleaned_items = []
    for item in extracted_data['boq_items']:
        # Skip invalid items
        if not item.get('item_name') and not item.get('description'):
            continue

        # Ensure numeric fields are valid
        item['quantity'] = float(item.get('quantity', 0) or 0)
        item['rate'] = float(item.get('rate', 0) or 0)
        item['amount'] = float(item.get('amount', 0) or 0)

        # Calculate amount if missing
        if item['amount'] == 0 and item['quantity'] > 0 and item['rate'] > 0:
            item['amount'] = item['quantity'] * item['rate']

        cleaned_items.append(item)

    extracted_data['boq_items'] = cleaned_items
    return extracted_data

def add_to_master_tables(item_name, description, work_type, materials_data, labour_data, created_by, overhead_percentage=None, overhead_amount=None, profit_margin_percentage=None, profit_margin_amount=None):
    """Add items, materials, and labour to master tables - PROPERLY AVOIDING DUPLICATES"""
    master_item_id = None
    master_material_ids = []
    master_labour_ids = []

    # Normalize item name for comparison (case-insensitive, trimmed)
    normalized_item_name = item_name.strip().lower()

    # Check for existing item with case-insensitive search
    master_item = MasterItem.query.filter(
        db.func.lower(MasterItem.item_name) == normalized_item_name
    ).first()

    if not master_item:
        # Create new item only if it doesn't exist
        master_item = MasterItem(
            item_name=item_name.strip(),  # Store with original casing but trimmed
            description=description,
            overhead_percentage=overhead_percentage,
            overhead_amount=overhead_amount,
            profit_margin_percentage=profit_margin_percentage,
            profit_margin_amount=profit_margin_amount,
            created_by=created_by
        )
        db.session.add(master_item)
        db.session.flush()
    else:
        # Update description and overhead/profit values
        if description:
            master_item.description = description

        # Always update overhead and profit values with latest calculations
        master_item.overhead_percentage = overhead_percentage
        master_item.overhead_amount = overhead_amount
        master_item.profit_margin_percentage = profit_margin_percentage
        master_item.profit_margin_amount = profit_margin_amount

        db.session.flush()
    master_item_id = master_item.item_id

    # Add to master materials - PROPERLY AVOIDING DUPLICATES
    for mat_data in materials_data:
        material_name = mat_data.get("material_name", "").strip()
        if not material_name:
            continue

        # Normalize material name for comparison
        normalized_material_name = material_name.lower()
        unit_price = float(mat_data.get("unit_price", 0.0))

        # Check for existing material with case-insensitive search
        master_material = MasterMaterial.query.filter(
            db.func.lower(MasterMaterial.material_name) == normalized_material_name
        ).first()

        if not master_material:
            # Create new material only if it doesn't exist
            master_material = MasterMaterial(
                material_name=material_name,  # Store with original casing
                item_id=master_item_id,
                default_unit=mat_data.get("unit", "nos"),
                current_market_price=unit_price,
                created_by=created_by
            )
            db.session.add(master_material)
            db.session.flush()
        else:
            # Update existing material ONLY if values have changed
            updated = False

            # Update item_id if not set
            if master_material.item_id != master_item_id:
                master_material.item_id = master_item_id
                updated = True

            # Update price only if significantly different (avoid minor float differences)
            if abs(master_material.current_market_price - unit_price) > 0.01:
                master_material.current_market_price = unit_price
                updated = True

            # Update unit if different
            new_unit = mat_data.get("unit", "nos")
            if master_material.default_unit != new_unit:
                master_material.default_unit = new_unit
                updated = True

            if updated:
                db.session.flush()
            else:
                log.debug(f"No changes for master material: {master_material.material_name}")

        master_material_ids.append(master_material.material_id)

    # Add to master labour - PROPERLY AVOIDING DUPLICATES
    for labour_data_item in labour_data:
        labour_role = labour_data_item.get("labour_role", "").strip()
        if not labour_role:
            continue

        # Normalize labour role for comparison
        normalized_labour_role = labour_role.lower()

        # Get labour values from data
        rate_per_hour = float(labour_data_item.get("rate_per_hour", 0.0))
        hours = float(labour_data_item.get("hours", 0.0))
        labour_amount = float(labour_data_item.get("amount", 0.0))

        # Calculate amount if not provided
        if labour_amount == 0 and rate_per_hour > 0 and hours > 0:
            labour_amount = rate_per_hour * hours

        # Check for existing labour with case-insensitive search
        master_labour = MasterLabour.query.filter(
            db.func.lower(MasterLabour.labour_role) == normalized_labour_role
        ).first()

        if not master_labour:
            # Create new labour only if it doesn't exist
            master_labour = MasterLabour(
                labour_role=labour_role,  # Store with original casing
                item_id=master_item_id,
                work_type=work_type,
                hours=hours,
                rate_per_hour=rate_per_hour,
                amount=labour_amount,
                created_by=created_by
            )
            db.session.add(master_labour)
            db.session.flush()
        else:
            # Update existing labour ONLY if values have changed
            updated = False

            # Update item_id if different
            if master_labour.item_id != master_item_id:
                master_labour.item_id = master_item_id
                updated = True

            # Update work_type if different
            if master_labour.work_type != work_type and work_type:
                master_labour.work_type = work_type
                updated = True

            # Update hours if significantly different
            if master_labour.hours != hours and hours > 0:
                master_labour.hours = hours
                updated = True

            # Update rate_per_hour if significantly different
            if master_labour.rate_per_hour is None or (rate_per_hour > 0 and abs(master_labour.rate_per_hour - rate_per_hour) > 0.01):
                master_labour.rate_per_hour = rate_per_hour
                updated = True

            # Update amount only if significantly different
            if master_labour.amount is None or (labour_amount > 0 and abs(master_labour.amount - labour_amount) > 0.01):
                master_labour.amount = labour_amount
                updated = True

            if updated:
                db.session.flush()
            else:
                log.debug(f"No changes for master labour: {master_labour.labour_role}")

        master_labour_ids.append(master_labour.labour_id)

    return master_item_id, master_material_ids, master_labour_ids

def process_extracted_items_to_boq(extracted_items, project_id, boq_name, created_by, file_info=None):
    """Convert extracted items to BOQ structure

    Structure:
    - Sub-items (materials) are stored in boq_material table
    - Labour details are stored in boq_labours table
    - Each category becomes an item in boq_items table
    """
    try:
        # Create BOQ
        boq = BOQ(
            project_id=project_id,
            boq_name=boq_name,
            status="Draft",
            created_by=created_by,
        )
        db.session.add(boq)
        db.session.flush()

        # Group extracted items by parent category
        items_by_category = {}
        for item_data in extracted_items:
            parent = item_data.get('parent_item', 'General')
            if parent not in items_by_category:
                items_by_category[parent] = []
            items_by_category[parent].append(item_data)

        # Process items and create JSON structure
        boq_items = []
        total_boq_cost = 0
        total_materials = 0
        total_labour = 0

        # Process each category (main item)
        for category_name, category_items in items_by_category.items():
            materials_data = []
            labour_data = []
            work_type = category_items[0].get('work_type', 'General') if category_items else 'General'

            materials_cost_total = 0
            labour_cost_total = 0
            overhead_percentage = 10.0
            profit_margin_percentage = 15.0

            # Process each sub-item (material) in this category
            for item_data in category_items:
                material_name = item_data.get('item_name', 'Unknown Material')
                qty = float(item_data.get('qty', 1.0) or 1.0)
                unit = item_data.get('unit', 'nos')
                rate = float(item_data.get('rate', 0.0) or 0.0)
                amount = float(item_data.get('amount', 0.0) or 0.0)

                # If amount not provided, calculate it
                if amount == 0 and qty > 0 and rate > 0:
                    amount = qty * rate

                # Add material
                materials_data.append({
                    "material_name": material_name,
                    "quantity": qty,
                    "unit": unit,
                    "unit_price": rate
                })
                materials_cost_total += amount

                # Check for labour data
                if 'labour' in item_data and item_data['labour']:
                    labour_info = item_data['labour']
                    labour_role = labour_info.get('role', 'Worker')
                    working_hours = float(labour_info.get('working_hours', 0.0))
                    labour_amount = float(labour_info.get('amount', 0.0))
                    rate_per_hour = float(labour_info.get('rate_per_hour', 0.0))

                    if working_hours > 0 and (labour_amount > 0 or rate_per_hour > 0):
                        labour_data.append({
                            "labour_role": labour_role,
                            "hours": working_hours,
                            "rate_per_hour": rate_per_hour,
                            "amount": labour_amount
                        })
                        labour_cost_total += labour_amount
                # Get overhead and profit percentages if provided
                if item_data.get('overhead_percentage'):
                    overhead_percentage = float(item_data.get('overhead_percentage', 10.0))
                if item_data.get('profit_margin_percentage'):
                    profit_margin_percentage = float(item_data.get('profit_margin_percentage', 15.0))

            # Calculate totals for this category/item
            base_cost = materials_cost_total + labour_cost_total
            overhead_amount = (base_cost * overhead_percentage) / 100
            profit_margin_amount = (base_cost * profit_margin_percentage) / 100
            total_cost = base_cost + overhead_amount
            selling_price = total_cost + profit_margin_amount
            # Add to master tables
            master_item_id, master_material_ids, master_labour_ids = add_to_master_tables(
                category_name,
                f"Category: {category_name}",
                work_type,
                materials_data,
                labour_data,
                created_by,
                overhead_percentage,
                overhead_amount,
                profit_margin_percentage,
                profit_margin_amount
            )

            # Process materials for BOQ details
            item_materials = []
            for i, mat_data in enumerate(materials_data):
                quantity = mat_data.get("quantity", 1.0)
                unit_price = mat_data.get("unit_price", 0.0)
                total_price = quantity * unit_price

                item_materials.append({
                    "master_material_id": master_material_ids[i] if i < len(master_material_ids) else None,
                    "material_name": mat_data.get("material_name"),
                    "quantity": quantity,
                    "unit": mat_data.get("unit", "nos"),
                    "unit_price": unit_price,
                    "total_price": total_price
                })

            # Process labour for BOQ details
            item_labour = []
            for i, labour_data_item in enumerate(labour_data):
                hours = labour_data_item.get("hours", 0.0)
                rate_per_hour = labour_data_item.get("rate_per_hour", 0.0)
                amount = labour_data_item.get("amount", 0.0)

                # Use the amount from Excel if provided, otherwise calculate
                total_cost_labour = amount if amount > 0 else (hours * rate_per_hour)

                item_labour.append({
                    "master_labour_id": master_labour_ids[i] if i < len(master_labour_ids) else None,
                    "labour_role": labour_data_item.get("labour_role"),
                    "hours": hours,
                    "rate_per_hour": rate_per_hour,
                    "total_cost": total_cost_labour
                })

            # Create item JSON structure with properly calculated values
            item_json = {
                "master_item_id": master_item_id,
                "item_name": category_name,
                "description": f"Category: {category_name}",
                "work_type": work_type,
                "base_cost": base_cost,
                "overhead_percentage": overhead_percentage,
                "overhead_amount": overhead_amount,
                "profit_margin_percentage": profit_margin_percentage,
                "profit_margin_amount": profit_margin_amount,
                "total_cost": total_cost,
                "selling_price": selling_price,
                "totalMaterialCost": materials_cost_total,
                "totalLabourCost": labour_cost_total,
                "actualItemCost": base_cost,
                "estimatedSellingPrice": selling_price,
                "materials": item_materials,
                "labour": item_labour
            }

            boq_items.append(item_json)
            total_boq_cost += selling_price
            total_materials += len(item_materials)
            total_labour += len(item_labour)

        # Create BOQ details JSON
        boq_details_json = {
            "boq_id": boq.boq_id,
            "items": boq_items,
            "summary": {
                "total_items": len(boq_items),
                "total_materials": total_materials,
                "total_labour": total_labour,
                "total_material_cost": sum(item["totalMaterialCost"] for item in boq_items),
                "total_labour_cost": sum(item["totalLabourCost"] for item in boq_items),
                "total_cost": total_boq_cost,
                "selling_price": total_boq_cost,
                "estimatedSellingPrice": total_boq_cost
            }
        }

        # Save BOQ details with file information
        boq_details = BOQDetails(
            boq_id=boq.boq_id,
            boq_details=boq_details_json,
            # Summary fields
            total_cost=total_boq_cost,
            total_items=len(boq_items),
            total_materials=total_materials,
            total_labour=total_labour,
            # Add file information if provided
            file_name=file_info.get('file_name') if file_info else None,
            created_by=created_by
        )
        db.session.add(boq_details)
        db.session.commit()

        return {
            "boq_id": boq.boq_id,
            "boq_name": boq.boq_name,
            "project_id": boq.project_id,
            "status": boq.status,
            "total_cost": total_boq_cost,
            "items_count": len(boq_items),
            "materials_count": total_materials,
            "labour_count": total_labour
        }

    except Exception as e:
        # Don't rollback here, let the calling function handle it
        log.error(f"Error in process_extracted_items_to_boq: {str(e)}")
        raise e

# REMOVED: create_or_get_project() - BOQ now requires existing project_id
# REMOVED: create_boq_record() - BOQ creation handled in process_extracted_items_to_boq()
# REMOVED: save_boq_items() - Items stored in BOQDetails.boq_details JSON

def format_boq_structure_for_storage(items, project_details, file_url=None):
    """Format BOQ items into hierarchical structure for BOQDetails storage"""
    sections = {}
    grand_total = 0
    item_lookup = {}  # Store items by their item_number for quick lookup

    # First pass: Create all items and add to lookup
    for item in items:
        work_type = item.get('work_type', 'General')

        if work_type not in sections:
            sections[work_type] = {
                'title': work_type,
                'items': [],
                'total_amount': 0
            }

        # Format item for storage
        formatted_item = {
            'item_number': item.get('item_number', ''),
            'item_name': item.get('item_name', item.get('item', item.get('sub_item', item.get('description', '')))),
            'description': item.get('description', ''),
            'quantity': float(item.get('qty', 0) or 0),
            'unit': item.get('unit', 'LS'),
            'rate': float(item.get('rate', 0) or 0),
            'amount': float(item.get('amount', 0) or 0),
            'sub_items': []  # Will be populated if this is a parent item
        }

        # Store in lookup
        item_lookup[item.get('item_number', '')] = formatted_item

        # If it's not a sub-item, add to the section
        if not item.get('is_sub_item', False):
            sections[work_type]['items'].append(formatted_item)
            # Don't add main item amounts to total yet - will be calculated based on sub-items

    # Second pass: Organize sub-items under their parents and calculate totals
    for item in items:
        if item.get('is_sub_item', False):
            parent_num = item.get('parent_item_number')

            if parent_num and parent_num in item_lookup:
                # Add this item as a sub-item to its parent
                formatted_sub = {
                    'item_number': item.get('item_number', ''),
                    'item_name': item.get('item_name', item.get('sub_item', item.get('description', ''))),
                    'description': item.get('description', ''),
                    'quantity': float(item.get('qty', 0) or 0),
                    'unit': item.get('unit', 'LS'),
                    'rate': float(item.get('rate', 0) or 0),
                    'amount': float(item.get('amount', 0) or 0)
                }
                item_lookup[parent_num]['sub_items'].append(formatted_sub)

                # Update totals - only count sub-item amounts
                work_type = item.get('work_type', 'General')
                if work_type in sections:
                    sections[work_type]['total_amount'] += float(item.get('amount', 0) or 0)
                grand_total += float(item.get('amount', 0) or 0)

    # Third pass: Add amounts for items that have no sub-items
    for section_data in sections.values():
        for item in section_data['items']:
            # If this item has no sub-items, add its amount to the total
            if not item.get('sub_items', []):
                section_data['total_amount'] += float(item.get('amount', 0) or 0)
                grand_total += float(item.get('amount', 0) or 0)

    # Convert sections dict to list
    sections_list = [
        {
            'section_id': chr(65 + i),  # A, B, C, etc.
            'title': section['title'],
            'total_amount': section['total_amount'],
            'items': section['items']
        }
        for i, section in enumerate(sections.values())
    ]

    return {
        'title': f"BOQ - {project_details.get('project_name', 'Unknown Project')}",
        'sections': sections_list,
        'grand_total': grand_total,
        'total_items': len(items),
        'file_url': file_url,
        'work_types': list(sections.keys())
    }

def upload_to_supabase(file_content, file_name, boq_id):
    """Upload file to Supabase storage"""
    if not supabase:
        return None

    try:
        bucket_name = 'file_upload'
        file_path = f"{boq_id}/{file_name}"

        # Try to create bucket if it doesn't exist
        try:
            # Check if bucket exists by trying to list files
            supabase.storage.from_(bucket_name).list()
        except Exception as bucket_error:
            # Bucket might not exist, try to create it
            log.warning(f"Bucket '{bucket_name}' not found, attempting to create...")
            try:
                # Try using the anon key as service role key
                supabase.storage.create_bucket(
                    bucket_name,
                    {'public': True}
                )
            except Exception as create_error:
                log.error(f"Could not create bucket: {create_error}")
        # Try to upload file
        try:
            response = supabase.storage.from_(bucket_name).upload(
                file_path,
                file_content,
                {"content-type": "application/octet-stream"}
            )

            # Get public URL
            public_url = supabase.storage.from_(bucket_name).get_public_url(file_path)
            return public_url

        except Exception as upload_error:
            log.error(f"Could not upload to Supabase: {upload_error}")
            # Return None but continue processing
            return None

    except Exception as e:
        log.error(f"Error with Supabase storage: {e}")
        return None

def upload_boq_file():
    """Upload and process BOQ file for a specific project
    Args:
        project_id: The project ID to associate this BOQ with
    Returns:
        JSON response with BOQ details and extraction results
    """
    try:
        project_id = request.form.get('project_id')
        # Validate project exists
        project = Project.query.filter_by(project_id=project_id, is_deleted=False).first()
        if not project:
            return jsonify({'error': f'Project with ID {project_id} not found'}), 404

        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': f'Invalid file type. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

        # Get user information from request
        user_info = {
            'user_id': request.form.get('user_id', 1),
            'username': request.form.get('username', 'system')
        }

        # Read file content
        file_content = file.read()

        # Check file size
        if len(file_content) > MAX_FILE_SIZE:
            return jsonify({'error': 'File size exceeds 50MB limit'}), 400

        # Save file temporarily
        temp_dir = 'temp_uploads'
        os.makedirs(temp_dir, exist_ok=True)

        filename = secure_filename(file.filename)
        temp_path = os.path.join(temp_dir, f"{uuid.uuid4()}_{filename}")

        with open(temp_path, 'wb') as f:
            f.write(file_content)

        try:
            # Extract text and tables based on file type
            file_ext = filename.rsplit('.', 1)[1].lower()
            boq_data = {}

            if file_ext == 'pdf':
                text, tables, boq_data = extract_text_from_pdf(temp_path)
            elif file_ext in ['xlsx', 'xls']:
                text, tables = extract_text_from_excel(temp_path)
            else:
                text = file_content.decode('utf-8', errors='ignore')
                tables = []
            extracted_data = extract_boq_data(text, tables, boq_data)
            extracted_data = validate_and_clean_data(extracted_data)
            project_name = extracted_data.get('project_details', {}).get('project_name', '')
            if project_name:
                boq_name = truncate_and_clean_name(project_name, max_length=95)
            else:
                boq_name = truncate_and_clean_name(f'BOQ - {filename}', max_length=95)

            created_by = user_info.get('username', 'system')

            # Prepare file information
            file_info = {
                'file_name': filename
            }

            # Process BOQ with file info
            boq_result = process_extracted_items_to_boq(
                extracted_data.get('boq_items', []),
                project_id,
                boq_name,
                created_by,
                file_info
            )

            # Upload file to Supabase (optional - continues even if it fails)
            file_url = upload_to_supabase(file_content, filename, boq_result['boq_id'])

            # Update BOQDetails with additional extraction information
            boq_detail = BOQDetails.query.filter_by(boq_id=boq_result['boq_id']).first()
            if boq_detail:
                # Store formatted BOQ structure in boq_details JSON
                boq_structure = format_boq_structure_for_storage(
                    extracted_data.get('boq_items', []),
                    extracted_data.get('project_details', {}),
                    file_url
                )

                # Update the boq_details JSON with additional structure info
                existing_details = boq_detail.boq_details or {}
                existing_details['boq_structure'] = boq_structure
                existing_details['project_details'] = extracted_data.get('project_details', {})
                existing_details['extraction_metrics'] = {
                    'processing_time_seconds': 0,
                    'pages_processed': len(tables),
                    'items_extracted': len(extracted_data.get('boq_items', [])),
                    'extraction_method': 'enhanced-rule-based'
                }
                boq_detail.boq_details = existing_details
                boq_detail.file_name = filename

                db.session.merge(boq_detail)
                db.session.commit()
            return jsonify({
                'success': True,
                'message': 'BOQ file processed and integrated successfully',
                'boq_detail_id': boq_detail.boq_detail_id if boq_detail else None,
                'boq_id': boq_result['boq_id'],
                'project_id': project_id,
                'items_extracted': len(extracted_data.get('boq_items', [])),
                'file_name': filename,
                'boq_details': {
                    'boq_name': boq_name,
                    'total_cost': boq_result['total_cost'],
                    'items_count': boq_result['items_count'],
                    'materials_count': boq_result['materials_count'],
                    'labour_count': boq_result['labour_count']
                }
            }), 201

        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)

    except Exception as e:
        # Rollback any pending transaction
        try:
            db.session.rollback()
        except:
            pass  # If no transaction to rollback, continue
        return jsonify({'error': f'Processing failed: {str(e)}'}), 500

def get_extraction_details(boq_detail_id):
    """Get BOQ details by ID (keeping function name for compatibility)"""
    try:
        boq_detail = BOQDetails.query.filter_by(
            boq_detail_id=boq_detail_id,
            is_deleted=False
        ).first()

        if not boq_detail:
            return jsonify({'error': 'BOQ details not found'}), 404

        return jsonify({
            'boq_detail_id': boq_detail.boq_detail_id,
            'boq_id': boq_detail.boq_id,
            'file_name': boq_detail.file_name,
            # 'file_size': boq_detail.file_size,
            # 'file_type': boq_detail.file_type,
            # 'file_url': boq_detail.file_url,
            # 'extraction_method': boq_detail.extraction_method,
            # 'extraction_status': boq_detail.extraction_status,
            'total_cost': boq_detail.total_cost,
            'total_items': boq_detail.total_items,
            'total_materials': boq_detail.total_materials,
            'total_labour': boq_detail.total_labour,
            'boq_details': boq_detail.boq_details,
            'created_at': boq_detail.created_at.isoformat() if boq_detail.created_at else None,
            'created_by': boq_detail.created_by
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500